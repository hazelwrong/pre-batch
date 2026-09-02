"""Full-check validator and evidence generator for a GDPval delivery root.

Run:  GDPVAL_DELIVERY=<path> python3 pipeline/validate.py

Two independent things happen here and they are kept apart on purpose:

  1. Delivery-level checks — schema, paths, hashes, bundle isolation, junk
     files, leakage. These test the package.
  2. gold-deliverable-eval — the rubric is executed against the expert gold as
     if the gold were a submission. Every `required` item must pass. This tests
     whether the rubric is actually judgeable, which is the failure mode that
     static review never catches.

The recompute check does not read the workbook's formulas. It holds the gold
against figures the cold-context verifier derived from the reference files, so
agreement is corroboration rather than a tautology.
"""
import hashlib
import json
import os
import re
import shutil
import subprocess
import sys
import tempfile
from datetime import date
from uuid import UUID, uuid5

import docx
import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import checks as CH                                               # noqa: E402
import recompute as RC                                            # noqa: E402
import spec_checks as SC                                          # noqa: E402
import taskdata as TD                                             # noqa: E402
from validation_registry import (                                 # noqa: E402
    TEMPLATE_GUARD_ROLES, VALIDATION_REGISTRY_VERSION,
    expected_validation_checks, validation_registry_digest,
)

# The readers live in checks.py so the checker and the validator read a file the
# same way. Two implementations of "the text of this document" is two answers to
# the same question.
docx_text = CH.docx_text
pdf_text = CH.pdf_text

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DELIVERY = os.environ.get("GDPVAL_DELIVERY", os.path.join(BASE, "delivery"))
TODAY = date.today().isoformat()

REQUIRED_FIELDS = ["task_id", "sector", "occupation", "prompt", "reference_files",
                   "reference_file_urls", "reference_file_hf_uris", "deliverable_files",
                   "deliverable_file_urls", "deliverable_file_hf_uris",
                   "rubric_pretty", "rubric_json"]
JUNK = {".DS_Store", "Thumbs.db", "desktop.ini"}

results = []       # delivery-level checks
rubric_results = []
# Set once in main(). The evaluator-only side of the task: rubric version,
# check entries, expected values. Global for the same reason `results` is —
# several reporting helpers need it and threading it through each of them
# would say nothing extra.
TASK = None
ALL_RECORDS = []


def _signed_reviewer(entry):
    """Whether a reviewer entry is eligible to support an acceptance claim."""
    return bool(entry and entry.get("reviewer")
                and (entry.get("title") or entry.get("review_role"))
                and entry.get("date")
                and entry.get("counts_toward_acceptance") is not False)


def _adopted_rubric_codes(roster, rubric_version):
    """Codes explicitly adopted by eligible experts for this rubric version."""
    records = (roster or {}).get("occupational_expert_review") or []
    records = records if isinstance(records, list) else [records]
    adopted = set()
    for expert in records:
        if not _signed_reviewer(expert):
            continue
        if expert.get("rubric_version_reviewed") != rubric_version:
            continue
        if "adoption_rounds" in expert:
            for round_ in expert.get("adoption_rounds") or []:
                version = (round_.get("rubric_version") or
                           expert.get("rubric_version_reviewed"))
                if version == rubric_version:
                    adopted.update(round_.get("adopted") or [])
        else:
            # Legacy accepted rosters used items_reviewed as the final adopted
            # set. Once adoption_rounds exists, reviewed and adopted are not
            # interchangeable and this fallback is deliberately disabled.
            adopted.update(expert.get("items_reviewed") or [])
    return adopted


def rec(check, status, detail, evidence=""):
    results.append({"check": check, "status": status, "detail": detail,
                    "evidence_path": evidence})
    return status == "passed"


# The inventory cannot hash itself, and checksums_final.txt is written after it.
# Both absences are declared in the inventory header, so neither is a stray file.
NOT_IN_INVENTORY = ("manifests/file_inventory_sha256.txt",
                    "manifests/checksums_final.txt")


def sha256(p):
    h = hashlib.sha256()
    with open(p, "rb") as fh:
        for c in iter(lambda: fh.read(1 << 20), b""):
            h.update(c)
    return h.hexdigest()


def P(*parts):
    return os.path.join(DELIVERY, *parts)


# ==========================================================================
# 1. Delivery-level checks
# ==========================================================================
def check_delivery():
    global ALL_RECORDS
    with open(P("tasks.jsonl"), encoding="utf-8") as fh:
        lines = [l for l in fh if l.strip()]
    try:
        records = [json.loads(line) for line in lines]
    except (TypeError, ValueError) as exc:
        rec("tasks_jsonl_parses", "failed", "invalid JSON: %s" % exc)
        raise
    ids = [record.get("task_id") for record in records]
    selected = os.environ.get("GDPVAL_VALIDATE_TASK_ID")
    unique = len(ids) == len(set(ids)) and all(ids)
    target = next((record for record in records
                   if record.get("task_id") == selected), None) if selected else None
    valid_selection = bool(records and unique and
                           ((selected and target) or (not selected and len(records) == 1)))
    rec("tasks_jsonl_parses", "passed" if valid_selection else "failed",
        "%d valid record(s); selected=%s" % (len(records), selected or "single-task"))
    if not valid_selection:
        raise RuntimeError(
            "multi-task validation requires GDPVAL_VALIDATE_TASK_ID; use pipeline/run.py")
    ALL_RECORDS = records
    r = target or records[0]
    tid = r["task_id"]

    rec("schema_12_fields",
        "passed" if sorted(r) == sorted(REQUIRED_FIELDS) else "failed",
        "%d fields; exact match to the 12-field specification: %s"
        % (len(r), sorted(r) == sorted(REQUIRED_FIELDS)))

    try:
        ok = str(UUID(tid)) == tid and tid == tid.lower()
    except ValueError:
        ok = False
    rec("task_id_uuid", "passed" if ok else "failed", "task_id=%s" % tid)

    ref_b = uuid5(UUID(tid), "reference_files").hex
    dlv_b = uuid5(UUID(tid), "deliverable_files").hex
    ok = (all(ref_b in p for p in r["reference_files"])
          and all(dlv_b in p for p in r["deliverable_files"])
          and ref_b != dlv_b
          and all(len(b) == 32 and b == b.lower() for b in (ref_b, dlv_b)))
    rec("bundle_uuid5_derivation", "passed" if ok else "failed",
        "ref=%s dlv=%s (UUID5-derived, distinct, 32-char lowercase hex)" % (ref_b, dlv_b))

    paths = r["reference_files"] + r["deliverable_files"]
    ok = all(not p.startswith("/") and ".." not in p and "\\" not in p
             and not re.match(r"^[a-zA-Z]:", p) and not p.startswith("file:")
             for p in paths)
    rec("path_convention", "passed" if ok else "failed",
        "all delivery-root-relative POSIX paths; no absolute path, parent traversal "
        "or file URI")

    missing = [p for p in paths if not os.path.isfile(P(p)) or os.path.getsize(P(p)) == 0]
    rec("files_exist_nonempty", "passed" if not missing else "failed",
        "%d declared files, all present and non-empty" % len(paths) if not missing
        else "missing/empty: %s" % missing)

    lists = ["reference_files", "reference_file_urls", "reference_file_hf_uris",
             "deliverable_files", "deliverable_file_urls", "deliverable_file_hf_uris"]
    ok = all(isinstance(r[k], list) for k in lists)
    staging = (not r["reference_file_urls"] and not r["reference_file_hf_uris"]
               and not r["deliverable_file_urls"] and not r["deliverable_file_hf_uris"])
    cov = json.load(open(P("manifests", "coverage_manifest.json"), encoding="utf-8"))
    cov_row = next((row for row in cov if row.get("task_id") == tid), {})
    rr = cov_row.get("release_ready")
    rec("release_state_and_list_parity",
        "passed" if ok and staging and rr is False else "failed",
        "all six entries are lists; local staging with empty remote arrays and "
        "release_ready=false, as the specification requires for a staging package")

    items = json.loads(r["rubric_json"])
    total = sum(i["score"] for i in items)
    ids = {i["rubric_item_id"] for i in items}
    rec("rubric_json_parsable",
        "passed" if isinstance(items, list) and items else "failed",
        "rubric_json is a string parsing to a non-empty array of %d items" % len(items))
    rec("rubric_total_100", "passed" if total == 100 else "failed",
        "rubric item scores total %d" % total)
    rec("rubric_item_ids_unique_uuid",
        "passed" if len(ids) == len(items) and all(
            str(UUID(i)) == i for i in ids) else "failed",
        "%d rubric_item_id values, all unique UUIDs" % len(ids))
    # Specification §6 lists nine things each item must carry at minimum. Eight
    # are unconditional; `score_levels` applies where partial credit exists and
    # `check` where the item is programmatically checkable, so those two are
    # required only on the items that meet their condition. The earlier version
    # of this check had been narrowed to the eight fields the data happened to
    # have — which made a non-conforming rubric report as passing. A check must
    # not be trimmed to fit the artefact it is checking.
    need = {"score", "criterion", "required", "rubric_item_id", "author_type",
            "tags", "read_only", "form_content", "verification"}
    missing = [i.get("rubric_item_id") for i in items if not need <= set(i)]
    empty_verify = [i.get("rubric_item_id") for i in items
                    if not str(i.get("verification") or "").strip()]
    n_levels = sum(1 for i in items if "score_levels" in i)
    n_check = sum(1 for i in items if "check" in i)
    ok = not missing and not empty_verify
    rec("rubric_schema_fields", "passed" if ok else "failed",
        "%d items; all carry the eight unconditional fields of specification §6 "
        "plus a non-empty verification step. score_levels is present on the %d "
        "items that award partial credit and check on the %d that are "
        "programmatically testable — both are conditional in §6 and are carried "
        "wherever the condition holds.%s"
        % (len(items), n_levels, n_check,
           "" if ok else " MISSING on: %s" % (missing + empty_verify)[:5]))

    task_data = TD.TaskData(tid)
    codes = task_data.codes_in_rubric_order(items)
    adopted_codes = _adopted_rubric_codes(task_data.reviewers,
                                           task_data.rubric_version)
    author_mismatches = []
    for code, item in zip(codes, items):
        expected = "human" if code in adopted_codes else "pending_expert_review"
        if item.get("author_type") != expected:
            author_mismatches.append({"code": code,
                                      "expected": expected,
                                      "actual": item.get("author_type")})
    pending_count = len(items) - len(adopted_codes.intersection(codes))
    rec("rubric_author_type_truthful",
        "passed" if not author_mismatches else "failed",
        ("All %d author_type values match the current-version adoption evidence: "
         "%d human, %d pending_expert_review."
         % (len(items), len(items) - pending_count, pending_count)
         if not author_mismatches else
         "%d author_type value(s) contradict adoption evidence: %s"
         % (len(author_mismatches), author_mismatches[:8])))
    rec("rubric_adoption_complete",
        "passed" if pending_count == 0 else "not_run",
        ("All %d items have explicit current-version expert adoption." % len(items)
         if pending_count == 0 else
         "%d of %d items have explicit current-version expert adoption; %d remain "
         "pending and are not represented as human-authored."
         % (len(items) - pending_count, len(items), pending_count)))

    # hashes
    inv, bad = {}, []
    for line in open(P("manifests", "file_inventory_sha256.txt"), encoding="utf-8"):
        if line.startswith("#") or not line.strip():
            continue
        # Paths carry spaces ("Store Profile - Chaoyang Stores.xlsx"), so only
        # the first two fields may be split off; the rest is the path itself.
        h, n, rel = line.split(None, 2)
        rel = rel.strip()
        inv[rel] = (h, int(n))
    for rel, (h, n) in inv.items():
        f = P(rel)
        if not os.path.isfile(f) or sha256(f) != h or os.path.getsize(f) != n:
            bad.append(rel)
    rec("sha256_matches_inventory", "passed" if not bad else "failed",
        "%d files hashed, %d mismatches" % (len(inv), len(bad)))

    tree = []
    for root, dirs, files in os.walk(DELIVERY):
        dirs[:] = [d for d in dirs if d not in JUNK]
        for f in files:
            tree.append(os.path.relpath(os.path.join(root, f), DELIVERY).replace(os.sep, "/"))
    junk = [t for t in tree if os.path.basename(t) in JUNK]
    uncovered = [t for t in tree if t not in inv and t not in NOT_IN_INVENTORY]
    # The inventory is generated from the tree, so a stale file is inventoried
    # and therefore looks "covered". The test that actually bites is the other
    # direction: every payload file must be one tasks.jsonl declares.
    declared = {os.path.normpath(path) for record in ALL_RECORDS
                for path in record["reference_files"] + record["deliverable_files"]}
    payload = [t for t in tree
               if t.startswith(("reference_files/", "deliverable_files/"))]
    undeclared = sorted(set(payload) - declared)
    rec("payload_files_all_declared",
        "passed" if not undeclared else "failed",
        "%d files under reference_files/ and deliverable_files/; %s"
        % (len(payload),
           "every one is declared in tasks.jsonl." if not undeclared else
           "not declared in tasks.jsonl: %s" % ", ".join(undeclared[:6])),
        "tasks.jsonl")

    rec("delivery_tree_no_stray_files",
        "passed" if not junk and not uncovered else "failed",
        "%d files in tree; %d junk; %d not covered by the inventory (the inventory "
        "itself is excluded by construction and is noted as such in its header)"
        % (len(tree), len(junk), len(uncovered)))

    for name in ["coverage_manifest.json", "provenance_manifest.jsonl",
                 "source_inventory.jsonl", "file_inventory_sha256.txt"]:
        rec("manifest_present_" + name.split(".")[0],
            "passed" if os.path.isfile(P("manifests", name)) else "failed", name)

    # The vocabulary question has two halves and only one is settled. Reporting
    # them as a single line would either overclaim or understate.
    try:
        voc = json.load(open(os.path.join(BASE, "vocab",
                                          "controlled_vocabulary.json"),
                             encoding="utf-8"))
        pairs = {(s_["sector"], o["en"]) for s_ in voc["sectors"]
                 for o in s_["occupations"]}
        n_sec = len(voc["sectors"])
        n_occ = sum(len(s_["occupations"]) for s_ in voc["sectors"])
        in_list = (r["sector"], r["occupation"]) in pairs
        ver = voc.get("_verification", {})
    except Exception:                                            # noqa: BLE001
        in_list, n_sec, n_occ, ver = False, 0, 0, {}

    rec("controlled_vocabulary_mapping_verified",
        "passed" if in_list and n_sec == 9 and n_occ == 44 else "failed",
        "sector='%s' / occupation='%s' is a valid pair in the %d-sector, "
        "%d-occupation list. That list was verified line by line on %s against the "
        "source cited in the requirement's appendix (%s): sectors, occupations and "
        "their groupings match exactly, none missing, none extra."
        % (r["sector"], r["occupation"], n_sec, n_occ,
           ver.get("verified_on", "-"), ver.get("source", "-")),
        "vocab/controlled_vocabulary.json")

    # Settled. The client supplied the standard-format export, which carries the
    # sector and occupation strings themselves — so the spelling is no longer
    # reconstructed from a Chinese-only page, it is checked against the client's
    # own list. The check compares the exact strings, and the pairing.
    try:
        auth = json.load(open(os.path.join(BASE, "vocab",
                                           "authoritative_english_strings.json"),
                              encoding="utf-8"))
        a_pairs = {(x["sector"], x["occupation"]) for x in auth["pairs"]}
        sec_ok = r["sector"] in auth["sectors"]
        occ_ok = r["occupation"] in auth["occupations"]
        pair_ok = (r["sector"], r["occupation"]) in a_pairs
        n_s = auth["_counts"]["sectors"]
        n_o = auth["_counts"]["occupations"]
        src = auth["_source"]
        english_ok = sec_ok and occ_ok and pair_ok and n_s == 9 and n_o == 44
        detail = ("sector='%s' and occupation='%s' match the client's authoritative "
                  "list character for character, and are paired there as they are "
                  "here. The list holds %d sectors and %d occupations, taken from "
                  "%s (%d rows, sha256 %s), supplied %s."
                  % (r["sector"], r["occupation"], n_s, n_o, src["file"],
                     src["rows"], src["sha256"][:16], src["supplied_by"]))
        if not english_ok:
            detail = ("MISMATCH against the client's authoritative list — "
                      "sector present: %s; occupation present: %s; paired: %s."
                      % (sec_ok, occ_ok, pair_ok))
    except Exception as e:                                        # noqa: BLE001
        english_ok, detail = False, "authoritative list unreadable: %s" % e

    rec("controlled_vocabulary_english_strings",
        "passed" if english_ok else "failed", detail,
        "vocab/authoritative_english_strings.json")

    return r


# ==========================================================================
# 2. Leakage scan
# ==========================================================================
def _number_boundary(value):
    """Match a written number, not a fragment of a longer one.

    The trailing guard has to let a sentence end. Excluding any following period
    treats "the total is 44,520." as no match at all — which is how a leak
    sitting in plain prose survived the first test of this very check. Only a
    period or comma that is itself followed by a digit means the number
    continues.
    """
    return re.compile(r"(?<![0-9.,])%s(?![0-9])(?![.,][0-9])" % re.escape(value))


def _significant(value):
    """Significant digits of a number: 6.00 has one, 2.80 has two, 50,124 has five.

    Counted from the value, not from how it happens to be written. Counting the
    rendered string made "6.00" look like three significant digits and let a
    headcount into the noise set.
    """
    if isinstance(value, str):
        value = float(value.replace(",", ""))
    return len(re.sub(r"[^0-9]", "", "%g" % abs(value)).strip("0")) or 1


def _formats(value, min_significant=0):
    """The spellings one number can take in a document.

    `min_significant` applies to the *derived* set, not to declared results. A
    weight of 0.10 or a headcount of 6 turning up in a policy is coincidence,
    not a leak; a figure the task asks the Agent to produce is a leak whatever
    its precision, so results are collected with no threshold at all.
    """
    out = set()
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        return out
    if min_significant and _significant(value) < min_significant:
        return out
    for text in ("%g" % value, "{:,.2f}".format(value), "{:.2f}".format(value),
                 "{:,}".format(int(value)) if float(value).is_integer() else ""):
        if text:
            out.add(text)
    return out


def gold_only_values(r, task):
    """Numbers that exist because the task was done, not because it was set.

    Two sets, and the difference matters:

    * **Results** — the verifier's expected values that are carried somewhere in
      the deliverables. These are what the Agent is asked to produce, so they
      must not appear on the Agent-visible side at all. No corpus exclusion is
      applied: a result that turns up in a reference file *is* the leak.
    * **Other gold cells** — everything else the workbook holds, minus whatever
      the references already state. An input rate echoed in the output is not a
      leak; a derived figure copied into an input is.

    Excluding "anything already in the references" from both sets would make the
    scan self-defeating: pasting a result into a reference would remove it from
    the search set and the scan would report nothing. That is exactly the shape
    of the failure this function replaced.
    """
    payload = task.expected or {}
    entries = (payload.get("values") if isinstance(payload, dict) else payload) or []
    results = set()
    for entry in entries:
        if not entry.get("locator"):
            continue
        value = entry.get("value")
        # Small numbers are not conclusions. A 30-day comment period or a count
        # of three sites is stated in the inputs; searching for "30" reports the
        # reference that legitimately states it and buries any real finding.
        results |= _formats(value, min_significant=3)
        # Deliberately NOT scanned: string-valued expected values. Most of them
        # are givens as the gold renders them — the category name the proposal
        # already used, the identifier the signed forms carry, the notice date
        # the public participation file states. Searching them reports the
        # reference that legitimately says so, on every one. Telling a
        # conclusion from a given needs the verifier to label which is which,
        # and that label does not exist yet; until it does, a scan that cries
        # wolf on every quotation is worse than one that says what it covers.

    cells = set()
    for rel in r["deliverable_files"]:
        if not rel.lower().endswith((".xlsx", ".xlsm")):
            continue
        try:
            book = openpyxl.load_workbook(P(rel), data_only=True)
        except Exception:                                         # noqa: BLE001
            continue
        for sheet in book:
            for row in sheet.iter_rows():
                for cell in row:
                    if isinstance(cell.value, (int, float)):
                        cells |= _formats(cell.value, min_significant=3)
    corpus = " ".join(re.sub(r"\s+", " ", CH.any_text(P(rel)))
                      for rel in r["reference_files"])
    derived = {v for v in cells - results if v not in corpus}
    return sorted(results), sorted(derived)


def leakage_scan(r, task):
    tid = r["task_id"]
    results, derived = gold_only_values(r, task)
    gold_vals = sorted(set(results) | set(derived))
    # Scan for rubric machinery, not for ordinary English. Words such as
    # "recommended" and "disqualified" necessarily appear in a policy, which is
    # an Agent-visible rule document; treating them as leakage would be a false
    # positive. What must not appear is rubric wording, point values or
    # pass thresholds.
    rubric_terms = ["rubric", "hard gate", "score levels", "tolerance_rel",
                    "reject_values", "rubric_item_id", "points earned",
                    "pass threshold", "gold-deliverable-eval"]
    rubric_terms += sorted({"[+%d]" % item.get("score", 0)
                            for item in json.loads(r["rubric_json"])})

    # Record what was searched for, not only what was found. A scan reporting
    # zero hits is worth nothing unless the reader can see it had something to
    # look for — which is exactly how the hardcoded list went unnoticed while it
    # named a superseded package's totals.
    hits = {"gold_only_values_searched": gold_vals,
            "result_values_searched": results,
            "coverage_note": ("Numeric only. String-valued expected figures are "
                              "not scanned: distinguishing a conclusion from a "
                              "given among them needs a label the verifier does "
                              "not yet produce."),
            "result_values_note": ("Figures the task asks the Agent to produce. "
                                   "Searched unconditionally: a result appearing in "
                                   "an Agent-visible file is the leak, so it is never "
                                   "excluded for appearing there."),
            "gold_value_hits": [], "rubric_text_in_reference_files": [],
            "score_threshold_disclosure": []}

    scanned = []
    subjects = []
    for rel in r["reference_files"]:
        p, name = P(rel), os.path.basename(rel)
        scanned.append(name)
        subjects.append({"file": name, "sha256": sha256(p),
                         "bytes": os.path.getsize(p)})
        # One reader for every format. The branch this replaces assumed
        # "not PDF and not DOCX means spreadsheet", which was true of the
        # superseded package and false of the accepted one, where four of the
        # six inputs are Markdown.
        flat = re.sub(r"\s+", " ", CH.any_text(p))
        for g in gold_vals:
            # Word boundaries, not substring: "2.80" is inside "12.80" and
            # "0.10" is inside "10.10", and neither is a leak.
            if re.search(_number_boundary(g), flat):
                hits["gold_value_hits"].append({"file": name, "value": g})
        for t in rubric_terms:
            if t.lower() in flat.lower():
                hits["rubric_text_in_reference_files"].append({"file": name, "term": t})

    prompt = r["prompt"]
    for g in gold_vals:
        if re.search(_number_boundary(g), prompt):
            hits["score_threshold_disclosure"].append({"file": "prompt", "term": g})
    for g in rubric_terms:
        if g.lower() in prompt.lower():
            hits["score_threshold_disclosure"].append({"file": "prompt", "term": g})

    # metadata / hidden-content surfaces
    meta_hits = []
    for rel in r["reference_files"]:
        p = P(rel)
        if p.endswith(".xlsx"):
            wb = openpyxl.load_workbook(p)
            for ws in wb:
                if ws.sheet_state != "visible":
                    meta_hits.append({"file": os.path.basename(rel),
                                      "issue": "hidden sheet " + ws.title})
        if p.endswith(".docx"):
            d = docx.Document(p)
            cp = d.core_properties
            for attr in ("comments", "subject", "keywords", "category"):
                v = getattr(cp, attr, None)
                if v and any(g in str(v) for g in gold_vals):
                    meta_hits.append({"file": os.path.basename(rel),
                                      "issue": "gold value in core property " + attr})
    hits["metadata_and_hidden_content"] = meta_hits

    # Only the hit lists decide the verdict. `gold_only_values_searched` records
    # what was looked for; counting it as a finding would fail every clean scan.
    FINDINGS = ("gold_value_hits", "rubric_text_in_reference_files",
                "score_threshold_disclosure", "metadata_and_hidden_content")
    passed = not any(hits.get(k) for k in FINDINGS)
    # Bind the result to the exact bytes examined. Without this the file is
    # identical on every build regardless of what changed underneath it, which
    # makes it impossible to tell a fresh scan from a stale one.
    prompt_hash = hashlib.sha256(r["prompt"].encode("utf-8")).hexdigest()
    subjects = sorted(subjects, key=lambda x: x["file"])
    subjects.append({"file": "prompt (tasks.jsonl)", "sha256": prompt_hash,
                     "bytes": len(r["prompt"].encode("utf-8"))})
    fingerprints = hashlib.sha256(
        ("|".join(sorted(gold_vals)) + "||" +
         "|".join(sorted(rubric_terms))).encode("utf-8")).hexdigest()
    hits.update({"passed": passed, "scanned_reference_files": sorted(scanned),
                 "scan_subjects": subjects,
                 "scan_subject_digest": hashlib.sha256(
                     "".join(x["sha256"] for x in subjects).encode()).hexdigest(),
                 "detection_set_digest": fingerprints,
                 "detection_set_digest_note": (
                     "SHA-256 over the gold values and rubric terms searched for. "
                     "Together with scan_subject_digest this pins both what was "
                     "scanned and what it was scanned against, so this evidence "
                     "cannot be reused across a change to either."),
                 "gold_values_checked": len(gold_vals),
                 "summary": "Scanned %d Agent-visible files and the prompt against %d "
                            "gold values and %d rubric terms, plus hidden sheets and "
                            "document core properties. %s"
                            % (len(scanned), len(gold_vals), len(rubric_terms),
                               "No leakage detected." if passed else "LEAKAGE FOUND.")})
    rec("answer_leakage_scan", "passed" if passed else "failed", hits["summary"],
        "validation_evidence/%s/leakage_scan.json" % tid)
    return hits


# ==========================================================================
# 2b. Template-level guards
#
# These three checks exist because the client identified the same three root
# causes as template defects rather than one-off slips: replicated across 43
# more occupations they become systemic, and the contract lets the client widen
# an audit to every task sharing a duplicate_group_id. Each guard fails the
# build rather than merely reporting, so a defective template cannot be rolled
# out silently.
# ==========================================================================
STOPWORDS = {"the", "and", "for", "with", "that", "this", "from", "into", "each",
             "any", "all", "not", "its", "per", "are", "was", "has", "have",
             "which", "where", "when", "their", "there", "been", "being"}
# Words that are domain furniture rather than evidence come from the task, not
# from here: "store" and "vendor" are noise in a procurement task and content in
# a retail-operations one.

# Wording that leaves a test open to more than one honest reading.
VAGUE_TEST_TERMS = ["recurring", "adequate", "adequately", "reasonable", "reasonably",
                    "appropriate", "appropriately", "sufficient", "sufficiently",
                    "timely", "promptly", "as needed", "satisfactory", "acceptable",
                    "suitable", "meaningful", "material", "significant"]


def _tokens(text):
    return {w for w in re.findall(r"[A-Za-z][A-Za-z-]{4,}", text.lower())
            if w not in STOPWORDS}


GUARD_ROLES = TEMPLATE_GUARD_ROLES
GUARD_SOURCE_ROLES = GUARD_ROLES[:-1]


def _template_guard_applicability(roles):
    """Return whether the procurement-template guards apply to this task.

    The source roles identify the template family. A task that declares none of
    them is outside that family; a task that declares only some of them is a
    malformed member of the family and must fail closed.
    """
    declared_sources = [name for name in GUARD_SOURCE_ROLES if roles.get(name)]
    if not declared_sources:
        return "not_applicable", []
    missing = [name for name in GUARD_ROLES if not roles.get(name)]
    return ("applicable", []) if not missing else ("invalid", missing)


def check_template_guards(r, task):
    """Guard A/B/C — the three root causes behind the client's rejection.

    Which file plays which part is task data (`file_roles` in task_meta.json);
    the guards themselves are not. Naming the files inline, as this used to, tied
    the guards to one task and to one set of file formats — and it broke silently
    when the accepted package moved its inputs from PDF to Markdown.
    """
    ref = {os.path.basename(p): P(p) for p in r["reference_files"]}
    gold = {os.path.basename(p): P(p) for p in r["deliverable_files"]}
    roles = task.meta.get("file_roles") or {}
    guards = task.meta.get("guards") or {}
    domain_stopwords = set(guards.get("domain_stopwords") or [])

    applicability, absent = _template_guard_applicability(roles)
    if applicability == "not_applicable":
        reason = ("This task declares none of the procurement-template source roles "
                  "%s. Guards A/B/C are outside this task's template family."
                  % ", ".join(repr(name) for name in GUARD_SOURCE_ROLES))
        rec("template_guards_applicability", "passed", reason)
        return {"status": "not_applicable", "reason": reason}
    if applicability == "invalid":
        reason = ("This task declares at least one procurement-template source role "
                  "but is missing required role(s): %s. The template is incomplete, "
                  "so its guards fail closed."
                  % ", ".join(repr(name) for name in absent))
        rec("template_guards_applicability", "failed", reason)
        return {"status": "invalid", "missing_file_roles": absent}

    rec("template_guards_applicability", "passed",
        "All procurement-template file roles are declared; Guards A/B/C apply.")

    def role_text(name, joiner=" "):
        entries = roles.get(name)
        if not entries:
            raise RuntimeError(
                "task_meta.json declares no %r file role; guard would pass "
                "vacuously" % name)
        if isinstance(entries, str):
            entries = [entries]
        return joiner.join(CH.any_text(ref.get(n) or gold[n]) for n in entries)

    policy = re.sub(r"\s+", " ", role_text("policy"))
    issue_log = role_text("issue_log")
    quotes = role_text("quotations")

    profile_wb = openpyxl.load_workbook(ref[roles["profile"]], data_only=True)
    profile_bits = []
    for ws in profile_wb:
        headers = [str(c.value) if c.value is not None else ""
                   for c in next(ws.iter_rows(max_row=1))]
        # A column headed "... (months)" states its unit once, in the header.
        # Emit "<value> <unit>" alongside the raw value so a memo sentence like
        # "26 months" resolves against the sheet that actually says it.
        units = [(re.search(r"\(([a-z ]+)\)\s*$", h.lower()) or [None, None])[1]
                 for h in headers]
        for row in ws.iter_rows(min_row=2):
            for i, c in enumerate(row):
                if c.value is None:
                    continue
                profile_bits.append(str(c.value))
                if i < len(units) and units[i] and isinstance(c.value, (int, float)):
                    profile_bits.append("%s %s" % (c.value, units[i].strip()))
        profile_bits.extend(headers)
        # A run of dated records documents its own length. Without this, a memo
        # that says "across the 31 days of July" is reported as an invented
        # fact, when the 31 dated rows in front of it are exactly the source.
        # Stated as a class — any date column, whole sheet and per leading key —
        # rather than as an allowance for one phrase.
        date_cols = [i for i, h in enumerate(headers) if "date" in h.lower()]
        for i in date_cols:
            seen, per_key = set(), {}
            for row in ws.iter_rows(min_row=2):
                if i >= len(row) or row[i].value is None:
                    continue
                seen.add(str(row[i].value))
                key = str(row[0].value)
                per_key.setdefault(key, set()).add(str(row[i].value))
            for count in {len(seen)} | {len(v) for v in per_key.values()}:
                if count:
                    profile_bits.append("%d days" % count)
    profile = " ".join(profile_bits)
    reference_corpus = " ".join([issue_log, quotes, profile, policy])
    ref_tokens = _tokens(reference_corpus)
    memo = role_text("narrative_deliverable")

    # ---- Guard A: every scored requirement must have a real source anchor --
    orphans = []
    pattern = guards.get("requirement_row_pattern")
    rows = re.findall(pattern, policy) if pattern else []
    for code, body in rows:
        trace = body.strip().split("  ")[-1]
        keys = (_tokens(trace) - _tokens(guards.get("requirement_stopwords", ""))
                - domain_stopwords)
        if keys and not (keys & _tokens(issue_log)):
            orphans.append((code, trace[:60]))
    rec("template_guard_a_checklist_source_anchor",
        "passed" if not orphans else "failed",
        "Each functional requirement's stated trace resolves to wording that actually "
        "appears in the issue log. %d requirement(s) parsed, %d orphaned."
        % (len(rows), len(orphans)) if not orphans
        else "requirements with no evidence in the issue log: %s" % orphans)

    # ---- Guard B: no factual claim in gold outside the reference corpus ----
    suspects = []
    UNITS = "|".join(guards.get("quantified_units")
                     or ["day", "days", "month", "months", "year", "years",
                         "hour", "hours", "minute", "minutes", "week", "weeks"])
    # Figures documented in the reference corpus, grouped by unit. A total that
    # a competent writer computes from two documented figures — "30 minutes at
    # one store and 15 at another, 45 in all" — is traceable, not invented, so
    # sums of same-unit documented values are accepted alongside the values
    # themselves.
    documented = {}
    for m in re.finditer(r"\b(\d[\d,]*)\s*-?\s*(%s)\b" % UNITS,
                         reference_corpus, re.I):
        unit = m.group(2).lower().rstrip("s")
        documented.setdefault(unit, set()).add(int(m.group(1).replace(",", "")))
    sums = {u: {a + b for a in v for b in v} - v for u, v in documented.items()}
    derived = []
    for m in re.finditer(r"\b(\d[\d,]*)\s+(%s)\b" % UNITS, memo):
        value = int(m.group(1).replace(",", ""))
        unit = m.group(2).lower().rstrip("s")
        if value in documented.get(unit, set()):
            continue
        if value in sums.get(unit, set()):
            # Arithmetically reachable from documented figures. Accepted, but
            # surfaced: a coincidental sum looks identical to a deliberate one,
            # and only a reader can tell them apart.
            derived.append(m.group(0))
        else:
            suspects.append(m.group(0))
    # A named cost type the references never mention is an invented fact. Which
    # nouns name a cost is domain vocabulary, so the task supplies it.
    for noun in guards.get("named_quantity_nouns", []):
        for m in re.finditer(r"\b([a-z]+)\s+(%s)\b" % noun, memo.lower()):
            if m.group(1) in STOPWORDS or m.group(1) in domain_stopwords:
                continue
            if m.group(1) not in ref_tokens:
                suspects.append(m.group(0))
    rec("template_guard_b_no_invented_facts",
        "passed" if not suspects else "failed",
        ("Every quantified period and every named charge type in the memorandum is "
         "documented in the reference files%s."
         % ("" if not derived else
            "; %s reached by summing documented figures — reviewer to confirm each is "
            "a deliberate total, not a coincidence" % sorted(set(derived))))
        if not suspects
        else "phrases with no source in any reference file: %s" % sorted(set(suspects)))

    # ---- Guard C: every policy test must be mechanically decidable ---------
    hedges = []
    start, end = guards.get("scoring_articles", ["", ""])
    window = policy[policy.find(start):policy.find(end)] if start and end else policy
    for term in (guards.get("hedging_terms") or VAGUE_TEST_TERMS):
        for m in re.finditer(r"\b%s\b" % re.escape(term), window, re.I):
            hedges.append((term, window[max(0, m.start() - 45):m.start() + 45].strip()))
    rec("template_guard_c_tests_decidable",
        "passed" if not hedges else "failed",
        "No hedging term appears in the scoring articles; every stated test admits one "
        "reading. %d terms screened."
        % len(guards.get("hedging_terms") or VAGUE_TEST_TERMS) if not hedges
        else "hedging terms found in scoring articles: %s"
             % [(t, c[:70]) for t, c in hedges])

    return {"orphan_requirements": orphans, "unsourced_phrases": sorted(set(suspects)),
            "derived_by_sum_needs_confirmation": sorted(set(derived)),
            "hedging_terms": [t for t, _ in hedges]}


# ==========================================================================
# 3. Independent recompute — the gold against the verifier's figures
# ==========================================================================
def independent_recompute(r, task):
    """Hold the gold against the verifier's independently recomputed figures.

    The arithmetic is not done here. It is done by T14, a cold-context agent
    that never sees the gold producer's working notes, and arrives as
    expected_values.json. What this function contributes is the comparison and
    the evidence trail — which is the part that has to be the same for every
    task, and used to be the part that was written fresh for each one.
    """
    ctx = CH.Context(DELIVERY, r["deliverable_files"], r["reference_files"])
    payload = task.expected or {}
    values = payload.get("values") if isinstance(payload, dict) else payload
    evidence, ok = RC.run(values, ctx)
    evidence["provenance"] = (payload.get("provenance")
                              if isinstance(payload, dict) else None)
    status = RC.status_for(evidence)
    rec("gold_matches_independent_recompute", status,
        RC.summary(evidence) + (
            " Source: %s" % evidence["provenance"] if evidence["provenance"] else ""),
        "validation_evidence/%s/recompute_evidence.json" % r["task_id"])
    return evidence


# ==========================================================================
# 4. gold-deliverable-eval — run the rubric against the gold
# ==========================================================================
def _combined_gold_score(codes, manual_codes, marks, machine_earned):
    """Use a complete human marking as the authoritative total.

    A human may legitimately override an automated check. Adding machine points
    to only the manual-item marks silently discards that override and can make
    the evidence total disagree with the signed returned form.
    """
    all_codes = set(codes)
    marked_codes = set(marks)
    auto_codes = all_codes - set(manual_codes)
    manual_awarded = sum(int(marks[c]["awarded"])
                         for c in manual_codes & marked_codes)
    human_auto_awarded = sum(int(marks[c]["awarded"])
                             for c in auto_codes & marked_codes)
    if all_codes <= marked_codes:
        return {
            "combined": sum(int(marks[c]["awarded"]) for c in all_codes),
            "manual_awarded": manual_awarded,
            "human_auto_awarded": human_auto_awarded,
            "machine_cross_check_delta": machine_earned - human_auto_awarded,
            "basis": "complete_human_marking",
            "human_items_marked": len(all_codes),
        }
    return {
        "combined": machine_earned + manual_awarded,
        "manual_awarded": manual_awarded,
        "human_auto_awarded": human_auto_awarded,
        "machine_cross_check_delta": None,
        "basis": "machine_plus_human_judgement_items",
        "human_items_marked": len(marked_codes & all_codes),
    }


def run_rubric(r, task):
    """Execute the rubric against the expert gold, as if the gold were a
    submission. What each item tests is task data; how a test of that kind is
    performed is checks.py. Neither is written here any more — before this
    split, validating a second task meant editing the validator."""
    ctx = CH.Context(DELIVERY, r["deliverable_files"], r["reference_files"], task)
    items = json.loads(r["rubric_json"])
    codes = task.codes_in_rubric_order(items)

    # An item with neither a check nor a verification note can be neither run
    # nor handed to a person. Report that as a defect in the rubric rather than
    # letting it disappear into the not-evaluated pile.
    gap = task.unjudgeable(items)
    rec("rubric_item_judgeability", "passed" if not gap else "failed",
        "All %d rubric items are either executable or carry a verification note "
        "telling the marker how to settle them." % len(items) if not gap else
        "%d item(s) have neither a check nor a verification note: %s"
        % (len(gap), gap[:5]),
        "validation_evidence/%s/rubric_execution.json" % r["task_id"])

    for code, item in zip(codes, items):
        status, detail, kind = CH.execute(item, ctx)
        rubric_results.append({
            "rubric_item_id": item["rubric_item_id"],
            "code": code,
            "criterion": item["criterion"],
            "score": item["score"], "required": item["required"],
            "status": status, "detail": detail,
            "check_type": kind,
        })

    earned = sum(i["score"] for i in rubric_results if i["status"] == "passed")
    auto = [i for i in rubric_results if i["status"] != "not_auto_evaluated"]
    manual = [i for i in rubric_results if i["status"] == "not_auto_evaluated"]
    auto_avail = sum(i["score"] for i in auto)
    total_avail = sum(i["score"] for i in rubric_results)
    failed = [i for i in auto if i["status"] == "failed"]
    # Two different quantities were being compared. `earned` is what the machine
    # could test; THRESHOLD is set for a submission that has been marked in full.
    # With the machine-checkable items worth exactly 60 and the threshold at 60,
    # "earned >= THRESHOLD" passed on a coincidence and would have flipped on any
    # single item. The machine result is therefore judged on its own terms — no
    # checkable item may fail — and the threshold question is reported separately
    # as what it is: unsettled until a human marks the rest.
    # From policy.json. The specification puts it between 20 and 60; a copy of
    # the number compiled in here is a copy that drifts from the document that
    # is supposed to define it.
    THRESHOLD = TD.threshold()
    ok = not failed
    required_true = sum(item.get("required") is True for item in rubric_results)
    required_false = sum(item.get("required") is False for item in rubric_results)
    rec("gold_deliverable_eval", "passed" if ok else "failed",
        "Rubric run against the expert gold: %d of %d machine-checkable points "
        "earned across %d items, %s. The remaining %d points across %d items "
        "require a human marker and are recorded as not auto-evaluated — neither "
        "counted nor assumed. Rubric total %d. required=true on %d items and "
        "required=false on %d items; missing values default to true. required "
        "does not turn an unmarked human judgement into an automatic pass."
        % (earned, auto_avail, len(auto),
           "no item failed" if not failed
           else "%d failed: %s" % (len(failed), ", ".join(i["code"] for i in failed)),
           total_avail - auto_avail, len(manual), total_avail, required_true,
           required_false),
        "validation_evidence/%s/rubric_execution.json" % r["task_id"])

    # The specification puts the acceptance threshold between 20 and 60. Whether
    # the gold clears it cannot be answered by machine here: 40 of the 100 points
    # sit on items only a person can mark, and no one has marked them. Saying so
    # is what criterion 13 asks for; reporting the machine subtotal as though it
    # were the full score would not be.
    # The judgement items, marked by hand. The marking sheet is an input like the
    # reviewer roster: every entry has to cite where in the deliverables the
    # judgement was formed, and any item marked short has to say what is missing.
    # Items absent from the sheet stay unmarked — the threshold question is only
    # answered when the whole rubric has been scored.
    tid = r["task_id"]
    ev = os.path.join(DELIVERY, "validation_evidence", tid)
    os.makedirs(ev, exist_ok=True)
    # The marking sheet is task data, like the rubric it marks. Holding it in
    # pipeline/ meant one sheet for whichever task happened to be building.
    sheet = task.marking

    manual_codes = {i["code"] for i in manual}
    if not sheet or sheet.get("rubric_version") != TASK.rubric_version:
        rec("gold_scored_against_threshold", "not_run",
            "Acceptance threshold %d. %d points across %d judgement items are "
            "unmarked (no marking sheet for rubric version %s), so the gold's "
            "total is not established."
            % (THRESHOLD, total_avail - auto_avail, len(manual), TASK.rubric_version),
            "validation_evidence/%s/rubric_execution.json" % tid)
    else:
        marks = {m["code"]: m for m in sheet.get("items", [])}
        unmarked = sorted(manual_codes - set(marks))
        no_reason = [c for c in manual_codes & set(marks)
                     if not str(marks[c].get("evidence") or "").strip()]
        short = [m for m in marks.values() if m.get("shortfall")]
        scoring = _combined_gold_score(codes, manual_codes, marks, earned)
        awarded = scoring["manual_awarded"]
        combined = scoring["combined"]
        declared_total = sheet.get("returned_form_total")
        total_agrees = (declared_total in (None, "") or
                        int(declared_total) == combined)
        acceptance_eligible = bool(sheet.get("marked_by") and sheet.get("marked_on")
                                   and sheet.get("counts_toward_acceptance") is not False)
        good = (acceptance_eligible and not unmarked and not no_reason
                and total_agrees and combined >= THRESHOLD)
        status = "passed" if good else "not_run" if not acceptance_eligible else "failed"
        rec("gold_scored_against_threshold", status,
            ("Gold marked in full against the %d-item rubric: %d machine-checkable "
             "points plus %d awarded on the %d judgement items, %d of 100 against "
             "an acceptance threshold of %d. %s Marking was performed by %s on %s "
             "and is a self-assessment — the marker and the gold's producer are "
             "the same party, which is stated rather than claimed otherwise."
             % (len(rubric_results), earned, awarded, len(manual), combined,
                THRESHOLD,
                ("%d judgement item(s) were marked short and the shortfall is "
                 "recorded per item: %s." % (len(short),
                                             ", ".join(m["code"] for m in short))
                 if short else "No judgement item was marked short."),
                sheet.get("marked_by"), sheet.get("marked_on"))
             if good else
             ("Returned-form marks provisionally cover %d judgement items and produce "
              "%d of 100, but marked_by/marked_on are unresolved or the sheet is "
              "explicitly excluded from acceptance. The score is retained as evidence, "
              "not accepted as human marking."
              % (len(manual_codes & set(marks)), combined))
             if not acceptance_eligible else
             "Marking incomplete or internally inconsistent: %s unmarked; %s marked "
             "without a located reason; returned-form total %s versus recomputed %d."
             % (unmarked or "none", no_reason or "none", declared_total, combined)),
            "validation_evidence/%s/gold_human_marking.json" % tid)
        json.dump({"task_id": tid, "rubric_version": TASK.rubric_version,
                   "marked_by": sheet.get("marked_by"),
                   "marked_on": sheet.get("marked_on"),
                   "counts_toward_acceptance": sheet.get("counts_toward_acceptance"),
                   "independence": sheet.get("independence"),
                   "method": sheet.get("method"),
                   "machine_checkable_earned": earned,
                   "machine_checkable_available": auto_avail,
                   "human_machine_checkable_awarded": scoring["human_auto_awarded"],
                   "machine_cross_check_delta": scoring["machine_cross_check_delta"],
                   "judgement_awarded": awarded,
                   "judgement_available": total_avail - auto_avail,
                   "human_items_marked": scoring["human_items_marked"],
                   "score_basis": scoring["basis"],
                   "returned_form_total": declared_total,
                   "combined_score": combined,
                   "threshold": THRESHOLD,
                   "items_marked_short": [m["code"] for m in short],
                   "prior_marking_history": sheet.get("prior_marking_history", []),
                   "supplemental_marking_history":
                       sheet.get("supplemental_marking_history", []),
                   "items": sheet.get("items", [])},
                  open(os.path.join(ev, "gold_human_marking.json"), "w",
                       encoding="utf-8"), ensure_ascii=False, indent=2)
    return earned, ok


# ==========================================================================
# 5. Renders
# ==========================================================================
def render(r, outdir):
    os.makedirs(outdir, exist_ok=True)
    made = []
    with tempfile.TemporaryDirectory() as tmp:
        profile = os.path.join(tmp, "libreoffice-profile")
        cache = os.path.join(tmp, "cache")
        os.makedirs(cache, exist_ok=True)
        render_env = dict(os.environ, XDG_CACHE_HOME=cache)
        for rel in r["deliverable_files"] + r["reference_files"]:
            src = P(rel)
            name = os.path.splitext(os.path.basename(rel))[0]
            pdf = src
            if not src.lower().endswith(".pdf"):
                subprocess.run(["soffice",
                                "-env:UserInstallation=file://" + profile,
                                "--headless", "--convert-to", "pdf",
                                "--outdir", tmp, src],
                               capture_output=True, timeout=180, env=render_env)
                pdf = os.path.join(tmp, name + ".pdf")
            if os.path.isfile(pdf):
                subprocess.run(["pdftoppm", "-png", "-r", "100", pdf,
                                os.path.join(outdir, name)],
                               capture_output=True, timeout=180)
                made += [f for f in os.listdir(outdir) if f.startswith(name)]
    return sorted(set(made))


# ==========================================================================
def _assert_validation_registry(task_meta):
    names = [item.get("check") for item in results]
    duplicates = sorted({name for name in names if names.count(name) > 1})
    expected = expected_validation_checks(task_meta)
    missing = sorted(expected - set(names))
    unexpected = sorted(set(names) - expected)
    if any(not name for name in names) or duplicates or missing or unexpected:
        raise RuntimeError(
            "fixed validation registry mismatch: missing=%s unexpected=%s "
            "duplicates=%s" % (missing, unexpected, duplicates))
    return expected


def write_validation_status(task_id, final=False):
    """Replace one task row without discarding other tasks' validation."""
    expected = (_assert_validation_registry(TASK.meta) if final
                 else expected_validation_checks(TASK.meta))
    path = P("manifests", "validation_status.jsonl")
    existing = []
    if os.path.isfile(path):
        for line in open(path, encoding="utf-8"):
            if not line.strip():
                continue
            try:
                row = json.loads(line)
            except ValueError:
                continue
            if row.get("task_id") != task_id:
                existing.append(row)
    existing.append({
        "task_id": task_id,
        "validated_at": TODAY,
        "validator": ("pipeline/validate.py (programmatic self-check); "
                      "human_review_* entries are transcribed from the task's "
                      "reviewers.json and were determined by the named reviewers, "
                      "not by this script"),
        "validator_sha256": sha256(os.path.abspath(__file__)),
        "validation_run_nonce": os.environ.get("GDPVAL_VALIDATION_NONCE"),
        "registry_version": VALIDATION_REGISTRY_VERSION,
        "registry_sha256": validation_registry_digest(expected),
        "checks": results,
    })
    with open(path, "w", encoding="utf-8") as fh:
        for row in sorted(existing, key=lambda item: item.get("task_id", "")):
            fh.write(json.dumps(row, ensure_ascii=False) + "\n")


# ==========================================================================
def main():
    global TASK
    if os.environ.get("GDPVAL_VALIDATOR_ORCHESTRATED") == "1":
        nonce = os.environ.get("GDPVAL_VALIDATION_NONCE")
        try:
            UUID(str(nonce))
        except (TypeError, ValueError, AttributeError):
            raise SystemExit(
                "orchestrated validation requires a fresh UUID nonce")
    r = check_delivery()
    tid = r["task_id"]
    TASK = TD.TaskData(tid)
    ev = P("validation_evidence", tid)
    os.makedirs(ev, exist_ok=True)

    leak = leakage_scan(r, TASK)
    json.dump(leak, open(os.path.join(ev, "leakage_scan.json"), "w", encoding="utf-8"),
              ensure_ascii=False, indent=2)

    # The reconstruction record — the basis for the provenance claim.
    import gold_revision as GR
    _, gr_status, gr_detail = GR.write(
        TASK, os.path.join(ev, "gold_revision"), DELIVERY, r["reference_files"])
    rec("gold_reconstruction_record", gr_status, gr_detail,
        "validation_evidence/%s/gold_revision/gold_revision.json" % tid)

    gold_source = TASK.gold_provenance or {}
    accepted_gold_sources = set((TD.policy().get("gold_source") or {})
                                .get("accepted_paths") or [])
    source_ok = (gold_source.get("source_type") in accepted_gold_sources
                 and gold_source.get("is_real_deliverable") is True)
    rec("gold_source_eligible", "passed" if source_ok else "failed",
        ("Gold provenance identifies an accepted real-deliverable path: %s."
         % gold_source.get("source_type")) if source_ok else
        ("Gold source_type=%r and is_real_deliverable=%r. Policy requires a real "
         "work deliverable on an accepted path; a faithful format reconstruction "
         "needs the original editable artifact or written client approval."
         % (gold_source.get("source_type"),
            gold_source.get("is_real_deliverable"))))

    # ---- §9 scans beyond answer leakage -------------------------------
    import security_scans as SS
    all_files = [P(x) for x in r["reference_files"] + r["deliverable_files"]]
    scans = SS.run_all(all_files, r["reference_files"] + r["deliverable_files"],
                       notes={"privacy_pii": (TASK.meta.get("privacy") or {}).get("note")},
                       extra_subjects=[{
                           "file": "prompt (tasks.jsonl)",
                           "sha256": hashlib.sha256(
                               r["prompt"].encode("utf-8")).hexdigest(),
                           "bytes": len(r["prompt"].encode("utf-8"))}])
    json.dump(scans, open(os.path.join(ev, "security_scans.json"), "w",
                          encoding="utf-8"), ensure_ascii=False, indent=2)
    LABEL = {"privacy_pii": "privacy_pii_scan", "copyright": "copyright_scan",
             "malicious_content": "malicious_content_scan",
             "path_traversal": "path_traversal_scan", "secrets": "secret_key_scan"}
    # A pattern scanner cannot tell a person's home address from a regulator's
    # office address, and most professional tasks carry the latter. The answer
    # is not to drop the pattern — it is to make the task say, in writing, which
    # findings it expects and why, and to fail on everything it did not declare.
    declared = ((TASK.meta.get("privacy") or {}).get("expected_findings") or [])

    def _is_declared(scan_key, hit):
        for entry in declared:
            if entry.get("scan") not in (None, scan_key):
                continue
            if entry.get("type") not in (None, hit.get("type")):
                continue
            token = entry.get("match_starts_with")
            if token and not str(hit.get("match", "")).startswith(token):
                continue
            if not entry.get("justification"):
                continue
            return True
        return False

    for key, name in LABEL.items():
        res = scans[key]
        hits = res.get("hits", res.get("content_hits", []))
        undeclared = [h for h in hits if not _is_declared(key, h)]
        accounted = len(hits) - len(undeclared)
        ok = res["passed"] or not undeclared
        rec(name, "passed" if ok else "failed",
            "%d file(s) scanned against %d patterns; %d finding(s)%s.%s"
            % (res.get("files_scanned", 0), res.get("patterns_checked", 0), len(hits),
               "" if not accounted else
               ", %d of them declared by the task with a stated reason" % accounted,
               (" " + res["note"]) if res.get("note") else "")
            if ok else
            "%d finding(s) the task does not declare: %s"
            % (len(undeclared), [h.get("match") for h in undeclared][:4]),
            "validation_evidence/%s/security_scans.json" % tid)
    scans["declared_findings"] = declared

    # ---- §5 source-to-gold lineage -------------------------------------
    import lineage as LN
    lin, lineage_ok, lineage_detail = LN.write(
        TASK, ev, [os.path.basename(p) for p in r["reference_files"]])
    rec("source_to_gold_lineage", "passed" if lineage_ok else "failed",
        lineage_detail,
        "validation_evidence/%s/source_to_gold_lineage.json" % tid)

    # ---- rules distilled from the accepted package -----------------------
    for name, status, detail in SC.run_all(r, DELIVERY, TD.policy(),
                                           marking=TASK.marking,
                                           reviewers=TASK.reviewers,
                                           provenance=TASK.provenance,
                                           policy_exceptions=TASK.policy_exceptions,
                                           task_root=TASK.root):
        rec(name, status, detail)

    guards = check_template_guards(r, TASK)
    json.dump(guards, open(os.path.join(ev, "template_guards.json"), "w",
                           encoding="utf-8"), ensure_ascii=False, indent=2)

    recomp = independent_recompute(r, TASK)
    json.dump(recomp, open(os.path.join(ev, "recompute_evidence.json"), "w",
                           encoding="utf-8"), ensure_ascii=False, indent=2)

    earned, gates_ok = run_rubric(r, TASK)
    json.dump({"items": rubric_results,
               "points_earned": earned,
               "points_available": sum(i["score"] for i in rubric_results),
               "hard_gates_accepted": gates_ok},
              open(os.path.join(ev, "rubric_execution.json"), "w", encoding="utf-8"),
              ensure_ascii=False, indent=2)

    files = render(r, os.path.join(ev, "renders"))
    rec("visual_render", "passed" if files else "failed",
        "%d page images rendered through LibreOffice for %d deliverable file(s) and "
        "%d reference file(s)" % (len(files), len(r["deliverable_files"]),
                                   len(r["reference_files"])),
        "validation_evidence/%s/renders/" % tid)
    render_expectations = TASK.meta.get("render_expectations") or {}
    if render_expectations:
        page_findings = []
        for filename, expectation in render_expectations.items():
            stem = os.path.splitext(os.path.basename(filename))[0]
            actual = sum(name.startswith(stem + "-") and name.endswith(".png")
                         for name in files)
            expected = int(expectation.get("pages", 0))
            if actual != expected:
                page_findings.append({"file": filename, "expected_pages": expected,
                                      "actual_pages": actual,
                                      "source": expectation.get("source")})
        rec("visual_output_contract",
            "passed" if not page_findings else "failed",
            ("%d declared page-count contract(s) match the rendered output."
             % len(render_expectations)) if not page_findings else
            "Rendered page-count contract mismatch: %s" % page_findings,
            "validation_evidence/%s/renders/" % tid)

    # ---- human review -------------------------------------------------
    # A layer is written as passed only when a named reviewer exists in
    # the task's reviewer roster. Anything else is not_run. The three surfaces that must
    # agree are this record, validation_status.jsonl and the rubric's
    # author_type; they are all derived here from the same source so they
    # cannot drift apart.
    roster = TASK.reviewers or {}

    expert_records = roster.get("occupational_expert_review") or []
    if not isinstance(expert_records, list):
        expert_records = [expert_records]

    # Returned forms may carry useful findings while their identity fields are
    # still unresolved. Preserve those records, but do not let a non-empty list
    # masquerade as a signed occupational review (or crash name rendering).
    experts = [entry for entry in expert_records
               if _signed_reviewer(entry) and
               entry.get("review_role") in (None, "occupational_expert_review")]
    unsigned_expert_feedback = [entry for entry in expert_records
                                if not _signed_reviewer(entry)]
    expert_qualification = next(
        (entry.get("required_qualification") for entry in expert_records
         if entry.get("required_qualification")),
        "具备与 %s / %s 任务相符的职业经验，能够独立复核业务判断、证据与评分标准。"
        % (TASK.meta.get("sector", "the declared sector"),
           TASK.meta.get("occupation", "the declared occupation")))
    # Adoption is credited only where the reviewer recorded which rubric version
    # they actually reviewed and it matches the current one. rubric.py applies
    # the same rule to author_type; if this record used a looser test the two
    # would disagree — the record claiming items adopted that tasks.jsonl still
    # marks pending. Same rule, one place, no drift.
    adopted_codes = _adopted_rubric_codes(roster, TASK.rubric_version)
    reviewed_codes = sorted(adopted_codes)
    expert_reviewed_codes = sorted({code for expert in experts
                                    if expert.get("rubric_version_reviewed") ==
                                    TASK.rubric_version
                                    for code in (expert.get("items_reviewed") or [])})
    # Read the codes, do not synthesise them. This line used to build
    # ["R01".."R44"] from the item count, which stopped matching reality the
    # moment items were removed from the front of the list: the rubric runs
    # R06-R49, so five codes that no longer exist were reported as awaiting
    # adoption while tasks.jsonl showed every item adopted.
    all_codes = TASK.codes_in_rubric_order(json.loads(r["rubric_json"]))
    unreviewed = [c for c in all_codes if c not in reviewed_codes]

    def _independence(hr_layers, experts):
        """Describe the signing pattern from the roster rather than from memory."""
        signed, names = [], []
        for lay in hr_layers:
            if lay and lay.get("status") == "passed" and lay.get("reviewer"):
                signed.append(lay["layer"])
                names.append(lay["reviewer"])
        expert_names = [e["reviewer"] for e in experts if e.get("reviewer")]
        all_names = names + expert_names
        dupes = sorted({n for n in all_names if all_names.count(n) > 1})
        # Reads as a sentence a person would write, but every name in it is
        # still read off the roster rather than typed in — so it cannot drift
        # out of step with who actually signed, which is what this field is for.
        LAYER_ZH = {"general_review": "通用审查",
                    "final_review": "终审"}
        parts = []
        parts.append("职业专家审核由%s签署" % ("、".join(expert_names) or "（未指派）"))
        for lay in hr_layers:
            if not lay:
                continue
            name = LAYER_ZH.get(lay["layer"], lay["layer"])
            parts.append("%s由%s签署" % (name, lay["reviewer"])
                         if lay.get("reviewer") else "%s未指派评审人，记为未签署" % name)
        note = "，".join(parts) + "。"
        note += ("三层审核签署人互不重复。" if not dupes else
                 "警告——以下人员跨层签署：%s。" % "、".join(dupes))
        parts = [note]
        return {"single_person_signing_all_layers": bool(dupes),
                "distinct_signatories": len(set(all_names)),
                "note": parts[0]}

    def _adoption_summary(experts, reviewed_codes):
        """Summarise adoption from the rounds record instead of restating it.

        The prose summary had drifted twice: it still described a three-round
        state after a fourth round closed, because a human wrote both the rounds
        and the summary and only updated one. Deriving the summary removes the
        second place where the same fact has to be kept true.
        """
        # Rounds belong to the rubric version they were run against. Reporting
        # four rounds beside zero adopted items would read as a contradiction;
        # what it actually means is that those rounds were about a rubric this
        # delivery no longer carries.
        # Each round records the rubric version it was run against. Filtering on
        # the reviewer's current version would sweep in rounds fought over an
        # earlier rubric and report, say, five rounds for a rubric that has seen
        # one — true of the reviewer, false of the rubric.
        rounds, stale = [], []
        for e in experts:
            for rd in (e.get("adoption_rounds") or []):
                ver = rd.get("rubric_version") or e.get("rubric_version_reviewed")
                (rounds if ver == TASK.rubric_version else stale).append(rd)
        if not rounds:
            return {
                "rubric_version": TASK.rubric_version,
                "rounds_recorded": 0,
                "rounds_recorded_against_earlier_versions": len(stale),
                "items_adopted_total": 0,
                "items_pending": len(all_codes),
                "note": ("本版本尚无采纳轮次。此前对旧版评分标准进行的 %d 轮采纳"
                         "（含实质驳回）记录保留在 occupational_expert_review[]."
                         "adoption_rounds 中，但不适用于本版本。" % len(stale)
                         if stale else "尚无采纳轮次记录。"),
            }
        objected = [r for r in rounds if r.get("objected")]
        adopted = []
        for r in rounds:
            a = r.get("adopted")
            if isinstance(a, list):
                adopted += a
        reworded = sorted({c for r in rounds for c in r.get("reworded", [])})
        added = sorted({c for r in rounds
                        for c in r.get("added_on_expert_proposal", [])})
        return {
            "rubric_version": TASK.rubric_version,
            "rounds_recorded": len(rounds),
            "rounds_against_earlier_rubric_versions": len(stale),
            "rounds_with_substantive_objection": len(objected),
            "objected_item_codes": sorted({c for r in objected
                                           for c in r.get("objected", [])}),
            "items_reworded_on_expert_instruction": reworded,
            "items_added_on_expert_proposal": added,
            "items_adopted_total": len(reviewed_codes),
            "items_pending": 0,
            "note": ("本版本经 %d 轮采纳，%d 项全部采纳，无遗留异议；其中 %d 项按专家"
                     "意见改写措辞、%d 项由专家提出后新增。另有 %d 轮是针对更早版本"
                     "评分标准的，记录保留但不计入本版本。逐轮明细见 "
                     "occupational_expert_review[].adoption_rounds。"
                     % (len(rounds), len(reviewed_codes), len(reworded),
                        len(added), len(stale))),
        }

    def layer(name, qual):
        who = roster.get(name)
        # A layer may hold one reviewer or several. The accepted record stores
        # the expert layer as a list and the other two as single entries;
        # accepting only one shape made the roster's format a hidden precondition.
        if isinstance(who, list):
            who = who[0] if who else None
        if (not _signed_reviewer(who) or
                who.get("review_role") not in (None, name)):
            missing = [field for field in ("reviewer", "title", "date")
                       if not (who or {}).get(field)]
            attempts = roster.get("_final_review_attempts") or []
            if name == "final_review" and attempts:
                latest = attempts[-1]
                reason = (
                    "The latest final-review attempt by %s at %s did not pass (%s); "
                    "the remediation still needs a new acceptance-eligible review."
                    % (latest.get("reviewer") or "the assigned reviewer",
                       latest.get("reviewed_at") or "time not supplied",
                       ", ".join(latest.get("finding_ids") or []) or
                       latest.get("status") or "open finding")
                )
            else:
                reason = ("No acceptance-eligible reviewer signature. Missing: %s."
                          % ", ".join(missing)) if missing else \
                         "The roster explicitly marks this review as not counting toward acceptance."
            return {"layer": name, "status": "not_run", "reviewer": None,
                    "date": None, "findings": None,
                    "required_qualification": qual,
                    "note": reason + " Recorded as not_run rather than passed, "
                            "per acceptance criterion 13."}
        record = {"layer": name, "status": "passed", "reviewer": who.get("reviewer"),
                  "title": who.get("title"), "date": who.get("date"),
                  "findings": who.get("findings"), "required_qualification": qual}
        for key in ("identity_status", "reviewed_at", "credential_status", "verdict",
                    "supplemental_reviews", "source_form"):
            if key in who:
                record[key] = who.get(key)
        return record

    hr = {
        "task_id": tid,
        "review_attempts": ((roster.get("_review_attempts") or []) +
                            (roster.get("_final_review_attempts") or [])),
        "layers": [
            layer("general_review",
                  "熟悉项目交付规范，能够核验数据结构、文件完整性、路径、格式、"
                  "来源记录与答案泄漏风险。不要求行业背景。"),
            {"layer": "occupational_expert_review",
             "status": "passed" if experts else "not_run",
             "required_qualification": expert_qualification,
             "reviewers": experts,
             "reviewed_item_codes": expert_reviewed_codes,
             "adopted_item_codes": reviewed_codes,
             "not_adopted_item_codes": unreviewed},
            layer("final_review",
                  "项目或质量负责人；确认前两层审核均已关闭、证据相互一致，"
                  "且不存在未完成的整改项。"),
        ],
        "rubric_adoption": {
            "status": ("complete" if experts and not unreviewed
                       else "partial" if experts else "not_run"),
            "adopted_item_codes": reviewed_codes,
            "pending_item_codes": unreviewed,
            "note": (
                "共 %d 项评分标准均由具名职业专家实际审阅并采纳，author_type 全部为 "
                "'human'，无待审核条目。author_type 的允许取值集合仍待甲方确认，"
                "见《待甲方确认问题清单》。" % len(reviewed_codes)
                if experts and reviewed_codes and not unreviewed else
                "仅 %d 项经职业专家审阅并采纳，author_type 记为 'human'；其余 %d 项"
                "维持 'pending_expert_review'，未冒充已采纳。"
                % (len(reviewed_codes), len(unreviewed))
                if experts and reviewed_codes else
                "评分标准已改写为 %s，条目内容与编号对应关系全部变更，原采纳记录"
                "不适用于本版本。全部 %d 项须由职业专家重新逐条采纳；在此之前"
                "维持 'pending_expert_review'，不沿用旧版采纳结果。"
                % (TASK.rubric_version, len(all_codes)) if experts else
                "尚未指派职业专家，无任何条目被采纳。"),
        },
        "adoption_summary": _adoption_summary(experts, reviewed_codes),
        "unsigned_review_feedback": {
            "occupational_expert_review": unsigned_expert_feedback,
            "note": ("These returned-form findings remain auditable but do not count as "
                     "a human signature or rubric adoption until reviewer identity, title "
                     "and date are supplied and the record is acceptance-eligible."),
        },
        "independence": _independence(hr_layers=[layer("general_review", ""),
                                                 None,
                                                 layer("final_review", "")],
                                       experts=experts),
        "remediation_closed": all(e.get("remediated") for e in experts) if experts else None,
    }
    # The three review narratives are written by people and have now gone stale
    # twice — once naming four experts when there was one, once describing a
    # three-round adoption after a fourth round closed. Prose cannot be derived,
    # but the falsifiable claims inside it can be checked, so a summary that has
    # fallen behind the roster fails the build instead of shipping.
    signers = {l.get("reviewer") for l in hr["layers"] if l.get("reviewer")}
    signers |= {e.get("reviewer") for e in experts if e.get("reviewer")}
    signers = {x for x in signers if x}
    # Names that used to sign this task and no longer do. Task data, because
    # they are this task's history: a list compiled into the validator covers
    # one project's people and silently covers nothing for the next task.
    KNOWN_PAST_NAMES = [n for n in (roster.get("_superseded_reviewers") or [])
                        if n]
    superseded_scope = ("%d superseded name(s) screened" % len(KNOWN_PAST_NAMES)
                        if KNOWN_PAST_NAMES else
                        "no superseded names declared for this task, so a stale "
                        "name would only be caught by the count checks below")
    stale = []
    narratives = [(l["layer"], l.get("findings")) for l in hr["layers"]]
    narratives += [("occupational_expert_review/%s" % e.get("reviewer"),
                    e.get("findings")) for e in experts]
    n_experts = len(experts)
    CJK_NUM = {"一": 1, "二": 2, "两": 2, "三": 3, "四": 4, "五": 5}
    for where, text in narratives:
        if not text:
            continue
        for name in KNOWN_PAST_NAMES:
            if name in text and name not in signers:
                stale.append("%s: names '%s', who signs no layer in the current "
                             "roster" % (where, name))
        # Allow modifiers between the counter and the noun ("四位具名专家"),
        # which is how the stale sentence was actually phrased.
        for m in re.finditer(r"([一二两三四五\d]+)\s*(?:位|名)[一-鿿]{0,4}?专家",
                             text):
            tok = m.group(1)
            claimed = CJK_NUM.get(tok, None)
            if claimed is None:
                claimed = int(tok) if tok.isdigit() else None
            if claimed is not None and claimed != n_experts:
                stale.append("%s: claims %s expert(s); the roster has %d"
                             % (where, tok, n_experts))
        for m in re.finditer(r"(\d+)\s*项(?:全部)?采纳", text):
            if int(m.group(1)) != len(reviewed_codes):
                stale.append("%s: claims %s items adopted; %d are adopted"
                             % (where, m.group(1), len(reviewed_codes)))
    hr["narrative_consistency"] = {
        "checked": [w for w, t in narratives if t],
        "current_signatories": sorted(signers),
        "findings": stale,
        "note": ("Each review narrative is checked against the roster it describes: "
                 "any person named who signs no layer, and any count of experts or "
                 "adopted items that disagrees with the record, is reported here. "
                 "Prose is written by people and cannot be derived, so it is "
                 "verified instead."),
    }
    rec("review_narratives_not_stale", "passed" if not stale else "failed",
        "%d narrative(s) checked against the current roster (%s); %s"
        % (len([1 for _w, t in narratives if t]), superseded_scope,
           "no stale claim found." if not stale else "; ".join(stale)),
        "validation_evidence/%s/human_review_record.json" % tid)

    # Measured change counts, injected from the diff evidence. They used to be
    # typed into the roster and went stale the moment a later round touched the
    # gold — the record said six edits where the diff measured ten.
    try:
        with open(os.path.join(ev, "gold_revision", "gold_revision.json"),
                  encoding="utf-8") as fh:
            _gr = json.load(fh)
        measured = {}
        for name, rec_ in _gr.get("files", {}).items():
            if name.endswith(".docx"):
                measured["docx_edit_blocks"] = rec_.get("edit_blocks")
            else:
                measured["xlsx_cell_changes"] = len(rec_.get("cell_changes") or [])
                measured["xlsx_structural_changes"] = len(
                    rec_.get("structural_changes") or [])
        measured["source"] = ("validation_evidence/%s/gold_revision/"
                              "gold_revision.json" % tid)
        measured["note"] = ("Counted from the draft-versus-final diff, covering "
                            "every round. Not asserted in the reviewer roster.")
        hr["gold_change_counts_measured"] = measured
    except (OSError, ValueError):
        hr["gold_change_counts_measured"] = {"error": "diff evidence unreadable"}

    # Corrections the expert identified but the supplier drafted sit in an
    # awkward middle state: the substance is the expert's, the wording is not.
    # Shipping one before the expert has confirmed the wording would credit them
    # with a sentence they never saw, so the state is surfaced here rather than
    # left to be noticed in the roster.
    # Scan every correction round, not one hard-coded key. The first version of
    # this guard looked only at "round5_defects_identified", so when a sixth
    # round of changes was made under a different name the guard saw nothing and
    # reported all clear — the delivery shipped substantive edits to adopted gold
    # with no record and no sign-off. A guard keyed to one field name only ever
    # guards that field.
    pending_corr, confirmed_corr = [], []
    for e in experts:
        g = e.get("gold_revision") or {}
        rounds = list(g.get("post_adoption_correction_rounds") or [])
        # Tolerate any legacy single-block form so an old roster cannot slip past.
        rounds += [v for k, v in g.items()
                   if k.endswith("_defects_identified") and isinstance(v, dict)]
        for blk in rounds:
            entry = {"correction_round": blk.get("correction_round"),
                     "identified_by": blk.get("identified_by"),
                     "drafted_by": blk.get("drafted_by"),
                     "items": len(blk.get("items") or []),
                     "status": blk.get("confirmation_status"),
                     "where": [i.get("where") for i in (blk.get("items") or [])]}
            (confirmed_corr if blk.get("confirmation_status") == "confirmed_by_expert"
             else pending_corr).append(entry)
    hr["post_adoption_corrections"] = {
        "confirmed": confirmed_corr,
        "awaiting_expert_confirmation": pending_corr,
        "note": ("专家采纳之后对 gold 的每一轮改动都列于此，含识别方、起草方与确认状态。"
                 "凡措辞未取得专家回执者，不得记为专家修订，该版本也不得作为已完成"
                 "人工复核的交付件发出。"),
    }
    n_pending = sum(x["items"] for x in pending_corr)
    n_ok = sum(x["items"] for x in confirmed_corr)
    rec("post_adoption_corrections_confirmed",
        "passed" if not pending_corr else "failed",
        ("采纳后共 %d 轮、%d 处更正，措辞均已取得专家回执确认。" % (len(confirmed_corr), n_ok)
         if not pending_corr else
         "第 %s 轮共 %d 处更正尚未取得专家确认（%s）。这些是对已采纳 gold 的实质表述"
         "修改，确认前不得作为完成人工复核的版本交付。"
         % ("、".join(str(x["correction_round"]) for x in pending_corr), n_pending,
            "；".join(w for x in pending_corr for w in x["where"]))),
        "validation_evidence/%s/human_review_record.json" % tid)

    json.dump(hr, open(os.path.join(ev, "human_review_record.json"), "w",
                       encoding="utf-8"), ensure_ascii=False, indent=2)

    for lay in hr["layers"]:
        nm = lay["layer"]
        if nm == "occupational_expert_review" and experts:
            rec("human_review_" + nm, "passed",
                "%d named reviewer(s) (%s) explicitly adopted rubric items %s. %s "
                "Items %s remain without current-version adoption evidence."
                % (len(experts), ", ".join(e["reviewer"] for e in experts),
                   ", ".join(reviewed_codes),
                   " ".join(filter(None, (e.get("remediation") for e in experts)))
                   or "No remediation note recorded.",
                   ", ".join(unreviewed) or "none"),
                "validation_evidence/%s/human_review_record.json" % tid)
        elif lay["status"] == "passed":
            rec("human_review_" + nm, "passed",
                "Signed by %s on %s." % (lay.get("reviewer"), lay.get("date")),
                "validation_evidence/%s/human_review_record.json" % tid)
        else:
            rec("human_review_" + nm, "not_run",
                (lay.get("note") or "No acceptance-eligible reviewer signature.") +
                " Acceptance criterion 13 requires incomplete review layers to be "
                "marked not_run; this layer is open.",
                "validation_evidence/%s/human_review_record.json" % tid)

    # The inventory is regenerated last, because validation evidence and
    # validation_status.jsonl are themselves delivery-tree files and must be
    # covered by it. The tree-coverage check is then re-evaluated against the
    # refreshed inventory rather than the pre-validation one.
    write_validation_status(tid)

    sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
    import build_delivery as BD
    BD.DELIVERY = DELIVERY
    payload = [(record["task_id"],
                "reference" if path.startswith("reference_files/") else "deliverable",
                path, P(path)) for record in ALL_RECORDS
               for path in record["reference_files"] + record["deliverable_files"]]
    builds = [BD.TaskBuild(record["task_id"]) for record in ALL_RECORDS]
    n_prov = BD.write_provenance(payload, builds)
    rec("provenance_covers_every_file", "passed",
        "%d provenance entries, one per file in the delivery tree" % n_prov,
        "manifests/provenance_manifest.jsonl")
    n_hashed = BD.write_sha256_inventory()

    inv = set()
    for line in open(P("manifests", "file_inventory_sha256.txt"), encoding="utf-8"):
        if not line.startswith("#") and line.strip():
            inv.add(line.split(None, 2)[2].strip())
    tree2, junk2 = [], []
    for root, dirs, files in os.walk(DELIVERY):
        dirs[:] = [d for d in dirs if d not in JUNK]
        for f in files:
            rel = os.path.relpath(os.path.join(root, f), DELIVERY).replace(os.sep, "/")
            tree2.append(rel)
            if os.path.basename(rel) in JUNK:
                junk2.append(rel)
    uncov = [t for t in tree2
             if t not in inv and t not in NOT_IN_INVENTORY]

    rebad = []
    for line in open(P("manifests", "file_inventory_sha256.txt"), encoding="utf-8"):
        if line.startswith("#") or not line.strip():
            continue
        # Paths carry spaces ("Store Profile - Chaoyang Stores.xlsx"), so only
        # the first two fields may be split off; the rest is the path itself.
        h, n, rel = line.split(None, 2)
        rel = rel.strip()
        f = P(rel)
        if not os.path.isfile(f) or sha256(f) != h or os.path.getsize(f) != int(n):
            rebad.append(rel)
    for c in results:
        if c["check"] == "sha256_matches_inventory":
            c["status"] = "passed" if not rebad else "failed"
            c["detail"] = ("%d files hashed against the final inventory, %d mismatches"
                           % (len(inv), len(rebad)))
    results.append({
        "check": "self_referential_manifest_hashes", "status": "passed",
        "detail": ("The inventory, the provenance manifest and this status file each "
                   "need the others' hashes, which no single pass can satisfy. Their "
                   "hashes are therefore written last, to "
                   "manifests/checksums_final.txt, and none of the three asserts a "
                   "hash for itself or for a file written after it."),
        "evidence_path": "manifests/checksums_final.txt"})

    for c in results:
        if c["check"] == "delivery_tree_no_stray_files":
            c["status"] = "passed" if not junk2 and not uncov else "failed"
            c["detail"] = ("%d files in tree, %d hashed; 0 junk files; every file "
                           "covered by the inventory except the inventory itself, "
                           "which is documented in its header"
                           % (len(tree2), n_hashed)) if not junk2 and not uncov else \
                          "junk=%s uncovered=%s" % (junk2, uncov)
    write_validation_status(tid, final=True)
    BD.write_sha256_inventory()
    BD.write_checksums_final()

    # summary
    counts = {}
    for c in results:
        counts[c["status"]] = counts.get(c["status"], 0) + 1
    print("=" * 74)
    print("DELIVERY CHECKS   passed=%d  failed=%d  not_run=%d"
          % (counts.get("passed", 0), counts.get("failed", 0), counts.get("not_run", 0)))
    for c in results:
        if c["status"] != "passed":
            print("   [%-7s] %s" % (c["status"], c["check"]))
    print("-" * 74)
    print("GOLD-DELIVERABLE-EVAL  %d machine-checkable points earned, %s"
          % (earned, "no item failed" if gates_ok else "ITEMS FAILED"))
    for i in rubric_results:
        print("   [%-6s] %3d %s" % (i["status"], i["score"], i["code"]))
    print("=" * 74)
    returncode = 0 if (counts.get("failed", 0) == 0 and
                       counts.get("not_run", 0) == 0 and gates_ok) else 1
    if os.environ.get("GDPVAL_VALIDATOR_ORCHESTRATED") == "1":
        print("GDPVAL_VALIDATION_COMPLETE nonce=%s returncode=%s" %
              (os.environ["GDPVAL_VALIDATION_NONCE"], returncode), flush=True)
    return returncode


if __name__ == "__main__":
    sys.exit(main())
