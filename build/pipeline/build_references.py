"""S-REF — build the Agent-visible reference files from a task's reference_spec.

The task designer decides what each input has to contain; this turns that into
files. It is the same job the old build_reference_files.py did, except that one
knew the six files of one task by name and could not build a seventh.

Two formats, per §1.3 of the build standard:

  * `.md` for policies, procedures, quotations and anything else a person would
    write as prose — clear structure, diffable, and a fraction of the tokens the
    same content costs as PDF.
  * `.xlsx` for registers, ledgers, schedules and detail data, which carry table
    semantics natively.
  * `.csv` / `.tsv` for flat extracts, and `.json` / `.xml` for system exports
    and payloads — the shapes a real system actually emits.

PDF is excluded outright: script-built PDFs share a give-away layout, which is
the fastest way to show a corpus was generated, and the same content costs
several times the tokens.

Spec shape (a list, one entry per file):

    {"filename": "<a name a real business document would carry>.md",
     "format": "md",
     "title": "<document title>",
     "meta": [["Issued by", "..."], ["Effective", "..."]],
     "blocks": [{"heading": "1. Scope"},
                {"text": "..."},
                {"list": ["...", "..."]},
                {"table": {"columns": ["...", "..."], "rows": [["...", "..."]]}}]}

    {"filename": "<register or ledger>.xlsx",
     "format": "xlsx",
     "sheets": [{"name": "<sheet>",
                 "columns": ["...", "...", "... (unit)"],
                 "rows": [["...", "...", 0]],
                 "widths": {"A": 12}, "number_formats": {"C": "0.00"},
                 "freeze": "A2"}]}

Output is deterministic: the same spec produces byte-identical files, because
every workbook goes through S-STRIP, which removes the document properties and
pins every zip member to one date.
"""
import json
import os
import re
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import officestrip                                                # noqa: E402

ALLOWED = ("md", "xlsx", "csv", "tsv", "xml", "json")
ENGINEERED_NAME = re.compile(r"^[a-z0-9]+(_[a-z0-9]+)+\.[a-z]+$")


def _validate_filename(name):
    """Require a single, relative filename rather than a path.

    Reference specs are copied into one staging directory.  Accepting a path
    here would let a malformed (or hostile) spec write through ``..`` or a
    platform-specific separator before any later validation gets a chance to
    run.
    """
    if not isinstance(name, str) or not name:
        raise ValueError("a reference_spec filename must be a non-empty string")
    if "\x00" in name:
        raise ValueError("%r contains a NUL byte" % name)
    if os.path.isabs(name) or re.match(r"^[A-Za-z]:[\\/]", name):
        raise ValueError("%r must be a relative filename, not an absolute path" % name)
    if "/" in name or "\\" in name:
        raise ValueError("%r must be a filename without path separators" % name)
    if name in (".", ".."):
        raise ValueError("%r is not a valid filename" % name)


def _safe_output_path(outdir, name):
    """Return a path whose resolved target remains under ``outdir``.

    This second check protects the write even when a file with the requested
    name already exists as a symlink pointing outside the staging tree.
    """
    root = os.path.realpath(os.path.abspath(os.fspath(outdir)))
    path = os.path.join(root, name)
    resolved = os.path.realpath(path)
    try:
        contained = os.path.commonpath((root, resolved)) == root
    except ValueError:  # Different drives on Windows, for example.
        contained = False
    if not contained:
        raise ValueError("reference output %r resolves outside outdir %r" %
                         (name, outdir))
    return path


def _cell(value):
    """Markdown table cells cannot carry a raw pipe or a newline."""
    text = "" if value is None else str(value)
    return text.replace("|", "\\|").replace("\n", " ")


def _meta_pairs(meta, filename):
    """The metadata header, however it was written.

    `for k, v in meta` silently unpacked a dict's *keys* when an author wrote
    `[{"key": ..., "value": ...}]`, so every line came out as the literal
    `- key: value` and the real metadata never reached the file. Nothing raised.
    Accepting one shape and mis-reading the others is worse than refusing them.
    """
    if isinstance(meta, dict):
        return list(meta.items())
    pairs = []
    for entry in meta:
        if isinstance(entry, dict):
            if "key" in entry and "value" in entry:
                pairs.append((entry["key"], entry["value"]))
                continue
            if "label" in entry and "value" in entry:
                pairs.append((entry["label"], entry["value"]))
                continue
            if len(entry) == 1:
                pairs.extend(entry.items())
                continue
            raise ValueError("%s: a meta entry has keys %s; use a pair, "
                             "{key, value}, {label, value} or a single-key object"
                             % (filename, sorted(entry)))
        if isinstance(entry, (list, tuple)) and len(entry) == 2:
            pairs.append(tuple(entry))
            continue
        raise ValueError("%s: a meta entry is %r; use a pair or an object"
                         % (filename, entry))
    return pairs


def render_markdown(spec):
    out = []
    if spec.get("title"):
        out.append("# %s" % spec["title"])
    if spec.get("meta"):
        pairs = _meta_pairs(spec["meta"], spec["filename"])
        out.append("\n".join("- %s: %s" % (k, v) for k, v in pairs))
    for block in spec.get("blocks", []):
        if "heading" in block:
            out.append("%s %s" % ("#" * int(block.get("level", 2)), block["heading"]))
        elif "text" in block:
            out.append(block["text"].strip())
        elif "list" in block:
            marker = block.get("marker", "-")
            out.append("\n".join("%s %s" % (marker, item) for item in block["list"]))
        elif "table" in block:
            table = block["table"]
            columns = table["columns"]
            lines = ["| %s |" % " | ".join(_cell(c) for c in columns),
                     "| %s |" % " | ".join("---" for _ in columns)]
            for row in table["rows"]:
                lines.append("| %s |" % " | ".join(_cell(v) for v in row))
            out.append("\n".join(lines))
        else:
            raise ValueError("unknown block in %s: %s" % (spec["filename"], block))
    return "\n\n".join(out) + "\n"


def write_markdown(spec, path):
    with open(path, "w", encoding="utf-8", newline="\n") as fh:
        fh.write(render_markdown(spec))
    return path


def _column_keyed(value, columns, what):
    """Accept the three shapes a person reasonably writes.

    A positional list, a dict keyed by column heading, or a dict keyed by
    column letter all say the same thing. Accepting only one of them turned a
    perfectly clear spec into an AttributeError from inside the writer, which
    is the wrong end of the pipeline to find out.
    """
    from openpyxl.utils import get_column_letter

    if not value:
        return {}
    heading_to_letter = {name: get_column_letter(n + 1)
                         for n, name in enumerate(columns or [])}
    if isinstance(value, list):
        if len(value) > len(columns or []):
            raise ValueError("%s lists %d entries for %d column(s)"
                             % (what, len(value), len(columns or [])))
        return {get_column_letter(n + 1): item for n, item in enumerate(value)
                if item is not None}
    if isinstance(value, dict):
        out = {}
        for key, item in value.items():
            letter = heading_to_letter.get(key, key)
            if not re.fullmatch(r"[A-Z]{1,3}", str(letter)):
                raise ValueError("%s names %r, which is neither a column heading "
                                 "nor a column letter" % (what, key))
            out[letter] = item
        return out
    raise ValueError("%s must be a list or an object, not %s"
                     % (what, type(value).__name__))


def write_workbook(spec, path, doc_date="2026-01-01"):
    import openpyxl
    from openpyxl.styles import Font

    book = openpyxl.Workbook()
    book.remove(book.active)
    for sheet_spec in spec["sheets"]:
        ws = book.create_sheet(sheet_spec["name"])
        columns = sheet_spec.get("columns")
        if columns:
            ws.append(list(columns))
            for cell in ws[1]:
                cell.font = Font(bold=True)
        for row in sheet_spec.get("rows", []):
            ws.append(list(row))
        widths = _column_keyed(sheet_spec.get("widths"), columns,
                               "%s / %s widths" % (spec["filename"], ws.title))
        for letter, width in widths.items():
            ws.column_dimensions[letter].width = width
        formats = _column_keyed(sheet_spec.get("number_formats"), columns,
                                "%s / %s number_formats" % (spec["filename"], ws.title))
        for letter, fmt in formats.items():
            for cell in ws[letter]:
                if cell.row > 1:
                    cell.number_format = fmt
        if sheet_spec.get("freeze"):
            ws.freeze_panes = sheet_spec["freeze"]
    book.save(path)
    # Every generated workbook goes through the strip, so the generator can
    # never be the thing that puts `Openpyxl 3.1.5` and a build clock into a
    # delivered file.
    officestrip.strip_and_pin(path, doc_date)
    return path


def write_delimited(spec, path, delimiter):
    """A flat extract, as a system would emit one."""
    import csv as _csv
    with open(path, "w", encoding="utf-8", newline="") as fh:
        writer = _csv.writer(fh, delimiter=delimiter, lineterminator="\n")
        if spec.get("columns"):
            writer.writerow(spec["columns"])
        for row in spec.get("rows", []):
            writer.writerow(["" if v is None else v for v in row])
    return path


def write_json(spec, path):
    """A system export. `data` is written as given; `rows` + `columns` are
    zipped into records, which is how an export of a table usually looks."""
    payload = spec.get("data")
    if payload is None:
        columns = spec.get("columns") or []
        payload = [dict(zip(columns, row)) for row in spec.get("rows", [])]
    with open(path, "w", encoding="utf-8") as fh:
        json.dump(payload, fh, ensure_ascii=False, indent=2)
        fh.write("\n")
    return path


def write_xml(spec, path):
    """An XML payload. Deliberately plain: elements and text, no attributes and
    no namespaces, because a reference file has to be readable by whoever opens
    it, not just parseable."""
    import xml.etree.ElementTree as ET

    def build(parent, value):
        if isinstance(value, dict):
            for key, item in value.items():
                build(ET.SubElement(parent, _tag(key)), item)
        elif isinstance(value, list):
            for item in value:
                build(ET.SubElement(parent, spec.get("item_tag", "item")), item)
        else:
            parent.text = "" if value is None else str(value)

    root = ET.Element(_tag(spec.get("root", "records")))
    payload = spec.get("data")
    if payload is None:
        columns = spec.get("columns") or []
        payload = [dict(zip(columns, row)) for row in spec.get("rows", [])]
    build(root, payload)
    ET.indent(root, space="  ")
    ET.ElementTree(root).write(path, encoding="utf-8", xml_declaration=True)
    return path


def _tag(name):
    """An element name a parser will accept, from a human-written heading."""
    tag = re.sub(r"[^A-Za-z0-9_.-]", "_", str(name)).strip("_") or "field"
    return tag if re.match(r"[A-Za-z_]", tag) else "f_" + tag


def validate(specs, policy=None):
    """Refuse a spec that would produce a package the standard rejects, before
    any file is written. Cheaper to fail here than to explain it downstream."""
    allowed = set((((policy or {}).get("reference_files") or {})
                   .get("allowed_formats")) or ALLOWED)
    forbidden = set((((policy or {}).get("reference_files") or {})
                     .get("forbidden_formats")) or ())
    if not specs:
        raise ValueError("reference_spec is empty; every task needs at least one "
                         "non-empty, openable reference file")
    seen = set()
    for spec in specs:
        name = spec.get("filename")
        _validate_filename(name)
        if name in seen:
            raise ValueError("two reference files share the basename %r" % name)
        seen.add(name)
        suffix = name.rsplit(".", 1)[-1].lower() if "." in name else ""
        if suffix in forbidden or suffix not in allowed:
            raise ValueError("%r is a .%s; only %s are accepted"
                             % (name, suffix, " and ".join(sorted(allowed))))
        declared = spec.get("format", suffix)
        if declared != suffix:
            raise ValueError("%r declares format %r" % (name, declared))
        if ENGINEERED_NAME.match(name):
            raise ValueError("%r reads as an engineered name; reference files carry "
                             "the names a real business document would" % name)
        if suffix == "xlsx":
            _validate_sheets(spec)
        if suffix in ("csv", "tsv", "json", "xml"):
            if not (spec.get("rows") or spec.get("data")):
                raise ValueError("%r needs rows or data" % name)
            for row in spec.get("rows") or []:
                if spec.get("columns") and len(row) > len(spec["columns"]):
                    raise ValueError("%s: a row has %d cells for %d column(s)"
                                     % (name, len(row), len(spec["columns"])))
        if spec.get("meta"):
            _meta_pairs(spec["meta"], name)
    return True


def _validate_sheets(spec):
    """Check a workbook spec before any file is written.

    A spec that fails here fails with a sentence naming the sheet and the field.
    A spec that fails inside the writer fails with an AttributeError and half a
    directory of files already on disk.
    """
    sheets = spec.get("sheets")
    if not isinstance(sheets, list) or not sheets:
        raise ValueError("%r needs a non-empty sheets list" % spec["filename"])
    for sheet in sheets:
        if not sheet.get("name"):
            raise ValueError("%r has a sheet with no name" % spec["filename"])
        columns = sheet.get("columns") or []
        for row in sheet.get("rows") or []:
            if columns and len(row) > len(columns):
                raise ValueError("%s / %s: a row has %d cells for %d column(s)"
                                 % (spec["filename"], sheet["name"], len(row),
                                    len(columns)))
        for field in ("widths", "number_formats"):
            _column_keyed(sheet.get(field), columns,
                          "%s / %s %s" % (spec["filename"], sheet["name"], field))


def build(specs, outdir, policy=None, doc_date="2026-01-01"):
    validate(specs, policy)
    os.makedirs(outdir, exist_ok=True)
    written = []
    for spec in specs:
        path = _safe_output_path(outdir, spec["filename"])
        suffix = spec["filename"].rsplit(".", 1)[-1].lower()
        if suffix == "md":
            written.append(write_markdown(spec, path))
        elif suffix == "xlsx":
            written.append(write_workbook(spec, path, doc_date))
        elif suffix in ("csv", "tsv"):
            written.append(write_delimited(spec, path, "," if suffix == "csv" else "\t"))
        elif suffix == "json":
            written.append(write_json(spec, path))
        elif suffix == "xml":
            written.append(write_xml(spec, path))
        else:
            raise ValueError("no writer for .%s" % suffix)
    return written


def main(argv):
    if len(argv) not in (2, 3):
        sys.exit("usage: build_references.py <reference_spec.json> [outdir]")
    with open(argv[1], encoding="utf-8") as fh:
        specs = json.load(fh)
    if isinstance(specs, dict):
        specs = specs.get("files") or specs.get("reference_spec") or []
    base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    outdir = argv[2] if len(argv) == 3 else os.path.join(base, "staging",
                                                         "reference_files")
    policy = {}
    try:
        import taskdata as TD
        policy = TD.policy()
    except Exception:                                             # noqa: BLE001
        pass
    for path in build(specs, outdir, policy):
        print("  %8d  %s" % (os.path.getsize(path), os.path.basename(path)))


if __name__ == "__main__":
    main(sys.argv)
